import json
import os
import pytest
from openpyxl import load_workbook


class TestLicenseAuditReport:
    """Test suite for software license audit task."""

    def test_json_report_exists(self):
        """Verify that the JSON license report file exists."""
        assert os.path.exists("/root/license_audit_report.json"), \
            "License audit report JSON file not found at /root/license_audit_report.json"

    def test_excel_report_exists(self):
        """Verify that the Excel license report file exists."""
        assert os.path.exists("/root/license_audit_report.xlsx"), \
            "License audit report Excel file not found at /root/license_audit_report.xlsx"

    def test_json_report_structure(self):
        """Verify that the JSON report has the correct structure."""
        with open("/root/license_audit_report.json", "r") as f:
            report = json.load(f)

        # Check required top-level keys
        required_keys = ["summary", "violations", "dependencies", "timestamp", "policy_version"]
        for key in required_keys:
            assert key in report, f"Missing required key '{key}' in JSON report"

        # Check summary structure and field types
        assert "total_dependencies" in report["summary"], \
            "Missing 'total_dependencies' in summary"
        assert isinstance(report["summary"]["total_dependencies"], int), \
            f"'total_dependencies' must be integer, got {type(report['summary']['total_dependencies'])}"

        assert "total_violations" in report["summary"], \
            "Missing 'total_violations' in summary"
        assert isinstance(report["summary"]["total_violations"], int), \
            f"'total_violations' must be integer, got {type(report['summary']['total_violations'])}"

        assert "violation_categories" in report["summary"], \
            "Missing 'violation_categories' in summary"

    def test_all_dependencies_analyzed(self):
        """Verify that all dependencies from input have been analyzed."""
        # Load input to get expected count
        with open("/root/project_dependencies.json", "r") as f:
            input_data = json.load(f)
        expected_count = len(input_data["dependencies"])

        with open("/root/license_audit_report.json", "r") as f:
            report = json.load(f)

        assert report["summary"]["total_dependencies"] == expected_count, \
            f"Expected {expected_count} dependencies, found {report['summary']['total_dependencies']}"
        assert len(report["dependencies"]) == expected_count, \
            f"Expected {expected_count} dependency entries, found {len(report['dependencies'])}"

    def test_runtime_violations_detected(self):
        """Verify that runtime license violations are correctly identified."""
        # Load input and policy to determine expected violations
        with open("/root/project_dependencies.json", "r") as f:
            input_data = json.load(f)
        with open("/root/license_policy.json", "r") as f:
            policy = json.load(f)

        with open("/root/license_audit_report.json", "r") as f:
            report = json.load(f)

        runtime_violations = [v for v in report["violations"] if v["dependency_type"] == "runtime"]

        # Verify each violation is actually a policy violation
        for violation in runtime_violations:
            dep = next((d for d in input_data["dependencies"]
                       if d["name"] == violation["name"]), None)
            assert dep is not None, f"Violation {violation['name']} not found in input"

            # Check if it's in restricted or prohibited lists
            is_violation = (violation["license"] in policy.get("restricted_licenses", {}).get("runtime", []) or
                          violation["license"] in policy.get("prohibited_licenses", {}).get("runtime", []))
            assert is_violation, f"{violation['name']} with {violation['license']} is not actually a violation"

    def test_development_violations_detected(self):
        """Verify that development license violations are correctly identified."""
        with open("/root/project_dependencies.json", "r") as f:
            input_data = json.load(f)
        with open("/root/license_policy.json", "r") as f:
            policy = json.load(f)
        with open("/root/license_audit_report.json", "r") as f:
            report = json.load(f)

        dev_violations = [v for v in report["violations"] if v["dependency_type"] == "development"]

        # Verify each dev violation is actually a policy violation
        for violation in dev_violations:
            dep = next((d for d in input_data["dependencies"]
                       if d["name"] == violation["name"] and d["type"] == "development"), None)
            assert dep is not None, f"Development violation {violation['name']} not found in input"

            # Check if it's in restricted or prohibited lists for development
            is_violation = (violation["license"] in policy.get("restricted_licenses", {}).get("development", []) or
                          violation["license"] in policy.get("prohibited_licenses", {}).get("development", []))

            assert is_violation, \
                f"{violation['name']} with {violation['license']} is not actually a development violation"

    def test_violation_details(self):
        """Verify that each violation contains exactly the required fields."""
        with open("/root/license_policy.json", "r") as f:
            policy = json.load(f)
        with open("/root/license_audit_report.json", "r") as f:
            report = json.load(f)

        # Define exactly the fields specified in instruction.md
        required_fields = {"name", "version", "license", "dependency_type", "violation_type"}

        for violation in report["violations"]:
            # Check all required fields are present
            for field in required_fields:
                assert field in violation, f"Violation missing required field '{field}'"

            # Check field types and values
            assert isinstance(violation["name"], str), "Violation 'name' must be string"
            assert isinstance(violation["version"], str), "Violation 'version' must be string"
            assert isinstance(violation["license"], str), "Violation 'license' must be string"
            assert violation["dependency_type"] in ["runtime", "development"], \
                f"Invalid dependency_type: {violation['dependency_type']}"
            assert violation["violation_type"] in ["restricted", "prohibited"], \
                f"Invalid violation_type: {violation['violation_type']}"

            # Check no extra fields beyond what's specified in instruction.md
            actual_fields = set(violation.keys())
            extra_fields = actual_fields - required_fields
            assert len(extra_fields) == 0, \
                f"Violation contains unexpected fields: {extra_fields}. Only {required_fields} are allowed."

    def test_complete_violation_detection(self):
        """Verify ALL real violations are detected and NO false positives exist."""
        with open("/root/project_dependencies.json", "r") as f:
            input_data = json.load(f)
        with open("/root/license_policy.json", "r") as f:
            policy = json.load(f)
        with open("/root/license_audit_report.json", "r") as f:
            report = json.load(f)

        # Build expected violations by manually checking each dependency against policy
        expected_violations = []
        expected_non_violations = []

        for dep in input_data["dependencies"]:
            dep_type = dep["type"]
            license_name = dep["license"]

            # Check if this dependency should be a violation
            is_restricted = license_name in policy.get("restricted_licenses", {}).get(dep_type, [])
            is_prohibited = license_name in policy.get("prohibited_licenses", {}).get(dep_type, [])

            if is_restricted or is_prohibited:
                violation_type = "prohibited" if is_prohibited else "restricted"
                expected_violations.append({
                    "name": dep["name"],
                    "version": dep["version"],
                    "license": dep["license"],
                    "dependency_type": dep_type,
                    "violation_type": violation_type
                })
            else:
                expected_non_violations.append({
                    "name": dep["name"],
                    "license": dep["license"],
                    "type": dep_type
                })

        # Convert reported violations to comparable format
        reported_violations = []
        for violation in report["violations"]:
            reported_violations.append({
                "name": violation["name"],
                "version": violation["version"],
                "license": violation["license"],
                "dependency_type": violation["dependency_type"],
                "violation_type": violation["violation_type"]
            })

        # Check ALL expected violations are reported (no missed violations)
        for expected in expected_violations:
            found = False
            for reported in reported_violations:
                if (expected["name"] == reported["name"] and
                    expected["version"] == reported["version"] and
                    expected["license"] == reported["license"] and
                    expected["dependency_type"] == reported["dependency_type"] and
                    expected["violation_type"] == reported["violation_type"]):
                    found = True
                    break
            assert found, f"Missing expected violation: {expected['name']} ({expected['license']}) should be {expected['violation_type']} for {expected['dependency_type']} dependencies"

        # Check NO false positives are reported (no non-violations in violations list)
        for reported in reported_violations:
            found_expected = False
            for expected in expected_violations:
                if (expected["name"] == reported["name"] and
                    expected["version"] == reported["version"] and
                    expected["license"] == reported["license"] and
                    expected["dependency_type"] == reported["dependency_type"] and
                    expected["violation_type"] == reported["violation_type"]):
                    found_expected = True
                    break
            assert found_expected, f"False positive violation reported: {reported['name']} ({reported['license']}) should NOT be a {reported['violation_type']} violation for {reported['dependency_type']} dependencies"

        # Verify counts match
        assert len(reported_violations) == len(expected_violations), \
            f"Violation count mismatch: expected {len(expected_violations)}, got {len(reported_violations)}"



    def test_excel_report_sheets(self):
        """Verify that the Excel report contains required sheets."""
        wb = load_workbook("/root/license_audit_report.xlsx")

        required_sheets = ["Summary", "Violations"]
        for sheet in required_sheets:
            assert sheet in wb.sheetnames, f"Missing required sheet '{sheet}' in Excel report"

    def test_excel_summary_sheet(self):
        """Verify the exact structure and ordering of the Excel Summary sheet."""
        wb = load_workbook("/root/license_audit_report.xlsx")
        summary_sheet = wb["Summary"]

        # Get all cells as a 2D array
        cells = [[cell.value for cell in row] for row in summary_sheet.iter_rows()]

        # Filter out completely empty rows
        non_empty_rows = [row for row in cells if any(cell is not None for cell in row)]

        # Verify exactly two columns structure (instruction says "two columns")
        for i, row in enumerate(non_empty_rows):
            # Each meaningful row should have exactly 2 columns (label and value)
            meaningful_cells = [cell for cell in row if cell is not None]
            if len(meaningful_cells) > 2:
                assert False, f"Row {i+1} has more than 2 meaningful cells: {meaningful_cells}. Expected exactly 2 columns."

        # Extract first column (labels) and second column (values) from non-empty rows
        labels = []
        values = []
        for row in non_empty_rows:
            if len(row) > 0 and row[0] is not None:
                labels.append(row[0])
                values.append(row[1] if len(row) > 1 else None)

        # Verify the exact sequence specified in instruction.md
        expected_labels_sequence = [
            "License Audit Report",
            "Report Generated",
            "Policy Version",
            "Summary Statistics",
            "Total Dependencies",
            "Total Violations",
            "Restricted License Violations",
            "Prohibited License Violations"
        ]

        # Verify exactly the specified labels are present, no more, no less
        expected_labels_set = set(expected_labels_sequence)
        actual_labels_set = set(labels)

        # Check for missing labels
        missing_labels = expected_labels_set - actual_labels_set
        assert len(missing_labels) == 0, \
            f"Missing required labels in Summary sheet: {missing_labels}"

        # Check for extra labels not specified in instruction.md
        extra_labels = actual_labels_set - expected_labels_set
        assert len(extra_labels) == 0, \
            f"Summary sheet contains extra labels not specified in instruction.md: {extra_labels}. Only these 8 labels are allowed: {expected_labels_sequence}"

        # Check that all required labels appear in order
        label_positions = {}
        for i, expected_label in enumerate(expected_labels_sequence):
            found_positions = [j for j, actual_label in enumerate(labels) if actual_label == expected_label]
            assert len(found_positions) > 0, \
                f"Required label '{expected_label}' not found in Summary sheet"
            label_positions[expected_label] = found_positions[0]

        # Verify ordering: each label should appear after the previous one
        for i in range(1, len(expected_labels_sequence)):
            prev_label = expected_labels_sequence[i-1]
            curr_label = expected_labels_sequence[i]
            assert label_positions[prev_label] < label_positions[curr_label], \
                f"Label '{curr_label}' should appear after '{prev_label}' but appears at position {label_positions[curr_label]} vs {label_positions[prev_label]}"

        # Verify that metric labels have corresponding values
        metric_labels = ["Total Dependencies", "Total Violations",
                        "Restricted License Violations", "Prohibited License Violations",
                        "Report Generated", "Policy Version"]

        for label in metric_labels:
            if label in label_positions:
                pos = label_positions[label]
                value = values[pos] if pos < len(values) else None
                assert value is not None, f"Label '{label}' missing corresponding value in column B"

                # Verify type constraints for numeric fields
                if label in ["Total Dependencies", "Total Violations",
                            "Restricted License Violations", "Prohibited License Violations"]:
                    is_valid = (isinstance(value, (int, float)) or
                               (isinstance(value, str) and value.startswith('=')))
                    assert is_valid, \
                        f"Label '{label}' should have numeric value or formula, got {type(value)}: {value}"

        # Verify data consistency with JSON report
        with open("/root/license_audit_report.json", "r") as f:
            json_report = json.load(f)

        # Verify exact label-to-value mapping for all labels
        expected_mappings = {
            "Report Generated": json_report["timestamp"],
            "Policy Version": json_report["policy_version"],
            "Total Dependencies": json_report["summary"]["total_dependencies"],
            "Total Violations": json_report["summary"]["total_violations"],
            "Restricted License Violations": json_report["summary"]["violation_categories"]["restricted"],
            "Prohibited License Violations": json_report["summary"]["violation_categories"]["prohibited"]
        }

        for label, expected_value in expected_mappings.items():
            assert label in label_positions, f"Required label '{label}' not found in Summary sheet"
            pos = label_positions[label]
            actual_value = values[pos] if pos < len(values) else None

            # Handle different value types and Excel formulas
            if isinstance(expected_value, int):
                # For numeric values, allow either the number itself or a formula
                if isinstance(actual_value, str) and actual_value.startswith('='):
                    # Formula detected - this is acceptable for numeric fields
                    pass
                elif isinstance(actual_value, (int, float)):
                    assert actual_value == expected_value, \
                        f"Label '{label}' has incorrect value: expected {expected_value}, got {actual_value}"
                else:
                    # Try converting string to int
                    try:
                        actual_int = int(actual_value)
                        assert actual_int == expected_value, \
                            f"Label '{label}' has incorrect value: expected {expected_value}, got {actual_int}"
                    except (ValueError, TypeError):
                        assert False, \
                            f"Label '{label}' should have numeric value {expected_value}, got non-numeric: {actual_value} (type: {type(actual_value)})"
            else:
                # For string values (timestamp, policy_version)
                assert str(actual_value) == str(expected_value), \
                    f"Label '{label}' has incorrect value: expected '{expected_value}', got '{actual_value}'"

    def test_excel_violations_sheet(self):
        """Verify the Violations sheet matches JSON violations exactly."""
        wb = load_workbook("/root/license_audit_report.xlsx")
        violations_sheet = wb["Violations"]

        # Load JSON violations for comparison
        with open("/root/license_audit_report.json", "r") as f:
            report = json.load(f)
        json_violations = report["violations"]

        # Check headers exactly match specification in instruction.md
        headers = [cell.value for cell in violations_sheet[1] if cell.value is not None]
        expected_headers = ["Package", "Version", "License", "Type", "Violation"]

        # Must have exactly the specified headers, no more, no less
        assert len(headers) == len(expected_headers), \
            f"Excel has {len(headers)} headers {headers}, expected exactly {len(expected_headers)} headers {expected_headers}"

        for i, expected_header in enumerate(expected_headers):
            assert i < len(headers), f"Missing header '{expected_header}' at position {i}"
            actual_header = headers[i]
            # Allow case-insensitive matching
            assert expected_header.lower() == actual_header.lower(), \
                f"Header at position {i}: expected '{expected_header}', got '{actual_header}'"

        # Find column indices
        header_map = {}
        for i, header in enumerate(headers):
            if header:
                if "package" in header.lower():
                    header_map["Package"] = i
                elif "version" in header.lower():
                    header_map["Version"] = i
                elif "license" in header.lower():
                    header_map["License"] = i
                elif "type" in header.lower():
                    header_map["Type"] = i
                elif "violation" in header.lower():
                    header_map["Violation"] = i

        # Verify all required columns are present
        required_cols = ["Package", "Version", "License", "Type", "Violation"]
        for col in required_cols:
            assert col in header_map, f"Column '{col}' not found in Excel headers"

        # Get data rows (skip header row)
        excel_violations = []
        for row_idx, row in enumerate(violations_sheet.iter_rows(min_row=2), 2):
            # Skip empty rows
            if all(cell.value is None for cell in row):
                continue

            violation_data = {}
            for col_name, col_idx in header_map.items():
                if col_idx < len(row):
                    violation_data[col_name] = row[col_idx].value
                else:
                    violation_data[col_name] = None

            # Only add if the row has meaningful data
            if any(violation_data[col] is not None for col in required_cols):
                excel_violations.append(violation_data)

        # Cross-validate: Excel violations should match JSON violations exactly
        assert len(excel_violations) == len(json_violations), \
            f"Excel has {len(excel_violations)} violation rows, JSON has {len(json_violations)}"

        # Check each violation matches
        for i, json_violation in enumerate(json_violations):
            # Find matching Excel violation by package name
            excel_violation = None
            for excel_v in excel_violations:
                if excel_v.get("Package") == json_violation["name"]:
                    excel_violation = excel_v
                    break

            assert excel_violation is not None, \
                f"JSON violation for '{json_violation['name']}' not found in Excel sheet"

            # Verify field mapping
            assert excel_violation["Package"] == json_violation["name"], \
                f"Package name mismatch: Excel '{excel_violation['Package']}' vs JSON '{json_violation['name']}'"
            assert excel_violation["Version"] == json_violation["version"], \
                f"Version mismatch for {json_violation['name']}"
            assert excel_violation["License"] == json_violation["license"], \
                f"License mismatch for {json_violation['name']}"
            assert excel_violation["Type"].lower() == json_violation["dependency_type"].lower(), \
                f"Type mismatch for {json_violation['name']}"
            assert excel_violation["Violation"].lower() == json_violation["violation_type"].lower(), \
                f"Violation type mismatch for {json_violation['name']}"


    def test_violation_count_accuracy(self):
        """Verify that the violation count matches the actual violations."""
        with open("/root/license_audit_report.json", "r") as f:
            report = json.load(f)

        actual_violations = len(report["violations"])
        reported_violations = report["summary"]["total_violations"]

        assert actual_violations == reported_violations, \
            f"Violation count mismatch: summary reports {reported_violations}, but found {actual_violations} violations"

    def test_license_categorization(self):
        """Verify that licenses are correctly categorized."""
        with open("/root/license_audit_report.json", "r") as f:
            report = json.load(f)

        categories = report["summary"]["violation_categories"]

        assert "restricted" in categories, "Missing 'restricted' category in summary"
        assert "prohibited" in categories, "Missing 'prohibited' category in summary"

        # Validate type constraints: counts should be integers
        assert isinstance(categories["restricted"], int), \
            f"Restricted count should be integer, got {type(categories['restricted'])}"
        assert isinstance(categories["prohibited"], int), \
            f"Prohibited count should be integer, got {type(categories['prohibited'])}"

        # Count actual violations by type
        restricted_count = len([v for v in report["violations"] if v["violation_type"] == "restricted"])
        prohibited_count = len([v for v in report["violations"] if v["violation_type"] == "prohibited"])

        assert categories["restricted"] == restricted_count, \
            f"Restricted count mismatch: reported {categories['restricted']}, actual {restricted_count}"
        assert categories["prohibited"] == prohibited_count, \
            f"Prohibited count mismatch: reported {categories['prohibited']}, actual {prohibited_count}"

    def test_dependency_schema_validation(self):
        """Verify that each dependency in the report follows the required schema exactly."""
        with open("/root/license_audit_report.json", "r") as f:
            report = json.load(f)

        # Define exactly the fields specified in instruction.md for dependencies
        required_fields = {"name", "version", "license", "type", "compliant"}

        for dep in report["dependencies"]:
            # Check all required fields are present
            for field in required_fields:
                assert field in dep, f"Missing required field '{field}' in dependency: {dep}"

            # Check no extra fields beyond what's specified in instruction.md
            actual_fields = set(dep.keys())
            extra_fields = actual_fields - required_fields
            assert len(extra_fields) == 0, \
                f"Dependency '{dep['name']}' contains unexpected fields: {extra_fields}. Only {required_fields} are allowed."

            # Check field types
            assert isinstance(dep["name"], str), f"'name' must be string in {dep['name']}"
            assert isinstance(dep["version"], str), f"'version' must be string in {dep['name']}"
            assert isinstance(dep["license"], str), f"'license' must be string in {dep['name']}"
            assert dep["type"] in ["runtime", "development"], f"Invalid type '{dep['type']}' in {dep['name']}"
            assert isinstance(dep["compliant"], bool), f"'compliant' must be bool in {dep['name']}"

    def test_timestamp_present(self):
        """Verify that the report includes a valid ISO8601 timestamp."""
        with open("/root/license_audit_report.json", "r") as f:
            report = json.load(f)

        assert "timestamp" in report, "Missing timestamp in report"
        assert report["timestamp"] != "", "Timestamp is empty"

        # Validate ISO8601 format
        import re
        iso8601_pattern = r'^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}(\.\d+)?([+-]\d{2}:\d{2}|Z)?$'
        assert re.match(iso8601_pattern, report["timestamp"]), \
            f"Timestamp '{report['timestamp']}' is not in valid ISO8601 format"


    def test_compliant_flag_consistency(self):
        """Verify that compliant flags are consistent with policy."""
        with open("/root/project_dependencies.json", "r") as f:
            input_data = json.load(f)
        with open("/root/license_policy.json", "r") as f:
            policy = json.load(f)
        with open("/root/license_audit_report.json", "r") as f:
            report = json.load(f)

        for dep in report["dependencies"]:
            # Find original input dependency
            input_dep = next(
                (d for d in input_data["dependencies"] if d["name"] == dep["name"]),
                None
            )
            assert input_dep is not None, f"Dependency {dep['name']} not found in input"

            dep_type = dep["type"]
            license_name = dep["license"]

            # Check if license violates policy
            is_violation = (
                license_name in policy.get("restricted_licenses", {}).get(dep_type, []) or
                license_name in policy.get("prohibited_licenses", {}).get(dep_type, [])
            )

            # If no violation, should be compliant; if violation, should NOT be compliant
            expected_compliant = not is_violation

            assert dep["compliant"] == expected_compliant, \
                f"Dependency {dep['name']}: compliant={dep['compliant']}, expected={expected_compliant} " \
                f"(violation={is_violation}, license={license_name})"


    def test_violation_type_accuracy(self):
        """Verify that violation_type correctly reflects license category."""
        with open("/root/license_policy.json", "r") as f:
            policy = json.load(f)
        with open("/root/license_audit_report.json", "r") as f:
            report = json.load(f)

        for violation in report["violations"]:
            license_name = violation["license"]
            dep_type = violation["dependency_type"]
            violation_type = violation["violation_type"]

            # Check if the license is in restricted or prohibited lists
            restricted_licenses = policy.get("restricted_licenses", {}).get(dep_type, [])
            prohibited_licenses = policy.get("prohibited_licenses", {}).get(dep_type, [])

            if license_name in restricted_licenses:
                assert violation_type == "restricted", \
                    f"License {license_name} is in restricted list but violation_type is {violation_type}"
            elif license_name in prohibited_licenses:
                assert violation_type == "prohibited", \
                    f"License {license_name} is in prohibited list but violation_type is {violation_type}"
            else:
                # This shouldn't happen - every violation should match a policy list
                assert False, f"Violation {violation['name']} has license {license_name} not found in policy lists"

    def test_policy_version_tracked(self):
        """Verify that the policy version is included in the report."""
        with open("/root/license_policy.json", "r") as f:
            policy = json.load(f)
        expected_version = policy["policy_version"]

        with open("/root/license_audit_report.json", "r") as f:
            report = json.load(f)

        assert "policy_version" in report, "Missing policy_version in report"
        assert report["policy_version"] == expected_version, \
            f"Expected policy version {expected_version}, found {report['policy_version']}"